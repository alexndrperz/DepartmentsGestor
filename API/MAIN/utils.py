import secrets, random, string,os, io, tempfile, pythoncom, time, multiprocessing
import subprocess
from django.http import HttpResponse, FileResponse
from openpyxl import load_workbook
from openpyxl.styles import   Border,Side, Font
from docx import Document
import docx2pdf, threading, docx2txt,mammoth
import win32com.client
import threading, pypandoc

class Services:
    def crear_token():
        token = secrets.token_urlsafe(32)
        token = f"{token}"
        return token
    
    def generate_code(length=6):
        letters = string.ascii_uppercase + string.ascii_lowercase + string.digits
        return ''.join(random.choice(letters) for _ in range(length))
    

    def generate_xlsx(data):
        border_style = Border(left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000'))
        

        workbook = load_workbook(os.path.abspath('dat_pl.xlsx'))
        worksheet = workbook['Hoja1']
        cell_dept = worksheet.cell(row=2, column=1)
        cell_dept.value =  str(cell_dept.value).replace('{{dep}}', data[0]['department']['name'])
        for row, item in enumerate(data, start=4):
            worksheet.cell(row=row, column=1, value=item['id']).border = border_style
            worksheet.cell(row=row, column=2, value=item['department']['name']).border = border_style
            worksheet.cell(row=row, column=3, value=item['producto']['name']).border = border_style
            worksheet.cell(row=row, column=4, value=item['createdAt']).border = border_style
            worksheet.cell(row=row, column=5, value=item['status']).border = border_style
        
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="archivo.xlsx"'
        excel_buffer.seek(0)
        response.write(excel_buffer.read())
        excel_buffer.close()
        return response
    







    def generate_pdf(data):
        template_path = os.path.abspath('plant_rep.docx')
        document = Document(template_path)
        pythoncom.CoInitialize()

        data = [
            {"nombre": "Ejemplo1", "valor": "Valor1"},
            {"nombre": "Ejemplo2", "valor": "Valor2"},
        ]

        table = document.tables[0]
        for item in data:
            row = table.add_row().cells
            row[0].text = item["nombre"]
        
        with open('plant_rep.docx', "rb") as docx_file:
            result = mammoth.convert_to_html(docx_file)
            text = result.value

        response = HttpResponse(text, content_type='text/html')
        response['Content-Disposition'] = 'inline; filename="archivo.html"'
        return response
